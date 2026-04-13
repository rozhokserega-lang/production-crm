"""Generate SQL upserts from Desktop/Соответствия.xlsx into repo migration folder."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = Path.home() / "Desktop" / "Соответствия.xlsx"
OUT_SQL = ROOT / "migration" / "generated_item_article_catalog_from_xlsx.sql"


def esc(s: str) -> str:
    return s.replace("'", "''")


def normalize_section(name: str) -> str:
    """Map Excel section headers to names used in public.section_catalog."""
    s = str(name or "").strip()
    aliases = {
        "Примьер черный": "Премьер черный",
        "Примьер белый": "Премьер белый",
        "Solito 1150 белый": "Solito 1150",
        "Solito 1150 черный": "Solito 1150",
        "Donini 806 белый": "Donini 806",
        "Donini 750 белый": "Donini 750",
    }
    return aliases.get(s, s)


def main() -> None:
    xlsx = DEFAULT_XLSX if DEFAULT_XLSX.exists() else Path(input("Path to Соответствия.xlsx: ").strip())
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    section: str | None = None
    order = 0
    rows: list[tuple[str, str, str, str, int]] = []

    for row in ws.iter_rows(values_only=True):
        a, b, c = (row + (None, None, None))[:3]
        if a is None and b and c is None:
            section = normalize_section(str(b).strip())
            order = 0
            continue
        if not a or not b or not c:
            continue
        order += 10
        art = str(a).strip()
        name = str(b).strip()
        color = str(c).strip()
        sec = normalize_section(section or "")
        rows.append((sec, art, name, color, order))

    wb.close()

    lines: list[str] = []
    lines.append("-- Generated from Соответствия.xlsx; re-run: python migration/scripts/gen_item_article_catalog_sql.py")
    lines.append("begin;")
    lines.append(
        "insert into public.item_article_map (article, item_name, source, section_name, table_color, sort_order)\nvalues"
    )
    vals = []
    for sec, art, name, color, so in rows:
        vals.append(
            f"  ('{esc(art)}', '{esc(name)}', 'xlsx_catalog', '{esc(sec)}', '{esc(color)}', {so})"
        )
    lines.append(",\n".join(vals))
    lines.append("on conflict (article) do update set")
    lines.append("  item_name = excluded.item_name,")
    lines.append("  section_name = excluded.section_name,")
    lines.append("  table_color = excluded.table_color,")
    lines.append("  sort_order = excluded.sort_order,")
    lines.append("  source = excluded.source,")
    lines.append("  updated_at = now();")
    lines.append("commit;")

    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUT_SQL} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
